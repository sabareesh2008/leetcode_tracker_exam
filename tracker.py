from __future__ import annotations

import csv
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, time as dt_time
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# SETTINGS
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

STUDENTS_FILE = BASE_DIR / "students.csv"
LIVE_CSV = BASE_DIR / "LiveData.csv"
HISTORY_CSV = BASE_DIR / "History.csv"
DAILY_ACTIVITY_CSV = BASE_DIR / "DailyActivity.csv"
STUDENTS_XLSX = BASE_DIR / "Students.xlsx"

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.getenv(
    "SUPABASE_SERVICE_ROLE_KEY",
    "",
)

LEETCODE_URL = "https://leetcode.com/graphql"

RECENT_SUBMISSION_LIMIT = 300
IST = ZoneInfo("Asia/Kolkata")

# Check up to 10 LeetCode profiles at the same time.
MAX_WORKERS = 20

ALLOWED_SECTIONS = (
    "ECE A",
    "ECE B",
    "ECE C",
    "ECE D",
    "ECE E",
    "ECE F",
)


LEETCODE_QUERY = """
query getUserProfile($username: String!, $limit: Int!) {
  matchedUser(username: $username) {
    username
    submitStatsGlobal {
      acSubmissionNum {
        difficulty
        count
        submissions
      }
    }
  }

  recentAcSubmissionList(
    username: $username,
    limit: $limit
  ) {
    title
    titleSlug
    timestamp
  }
}
"""


# ============================================================
# HELPERS
# ============================================================

def clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def empty_profile(status: str) -> dict[str, Any]:
    return {
        "total_solved": 0,
        "easy": 0,
        "medium": 0,
        "hard": 0,
        "submissions": 0,
        "solved_today": 0,
        "last_7_days": 0,
        "last_30_days": 0,
        "last_problem": "",
        "last_solved": "",
        "status": status,
        "recent_submissions": [],
    }


def get_stat(
    statistics: list[dict[str, Any]],
    difficulty: str,
    field: str = "count",
) -> int:
    for statistic in statistics:
        if statistic.get("difficulty") == difficulty:
            return int(statistic.get(field, 0) or 0)

    return 0


def unique_solved_count_since(
    recent_submissions: list[dict[str, Any]],
    start_time: datetime,
) -> int:
    solved = set()

    for submission in recent_submissions:
        timestamp = submission.get("timestamp")
        title_slug = submission.get("titleSlug")

        if not timestamp or not title_slug:
            continue

        submission_time = datetime.fromtimestamp(
            int(timestamp)
        )

        if submission_time >= start_time:
            solved.add(title_slug)

    return len(solved)


# ============================================================
# LEETCODE FETCH
# ============================================================

def fetch_leetcode(username: str) -> dict[str, Any]:
    if not username:
        return empty_profile("Username missing")

    request_body = {
        "operationName": "getUserProfile",
        "query": LEETCODE_QUERY,
        "variables": {
            "username": username,
            "limit": RECENT_SUBMISSION_LIMIT,
        },
    }

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Origin": "https://leetcode.com",
        "Referer": f"https://leetcode.com/u/{username}/",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/151.0.0.0 Safari/537.36"
        ),
    }

    try:
        response = requests.post(
            LEETCODE_URL,
            json=request_body,
            headers=headers,
            timeout=30,
        )

        if response.status_code != 200:
            return empty_profile(
                f"HTTP {response.status_code}"
            )

        try:
            response_data = response.json()
        except ValueError:
            return empty_profile(
                "Invalid JSON response"
            )

        if response_data.get("errors"):
            messages = [
                str(
                    error.get(
                        "message",
                        "GraphQL error",
                    )
                )
                for error in response_data["errors"]
            ]

            return empty_profile(
                " | ".join(messages)
            )

        data = response_data.get("data", {})
        matched_user = data.get("matchedUser")

        if matched_user is None:
            return empty_profile("User not found")

        statistics = (
            matched_user
            .get("submitStatsGlobal", {})
            .get("acSubmissionNum", [])
        )

        recent_submissions = (
            data.get("recentAcSubmissionList", [])
            or []
        )

        last_problem = ""
        last_solved = ""

        if recent_submissions:
            latest = recent_submissions[0]

            last_problem = clean(
                latest.get("title")
            )

            timestamp = latest.get("timestamp")

            if timestamp:
                last_solved = datetime.fromtimestamp(
                    int(timestamp),
                    tz=IST,
                ).strftime("%Y-%m-%d %H:%M:%S IST")

        now = datetime.now()

        today_start = datetime.combine(
            date.today(),
            datetime.min.time(),
        )

        seven_days_start = (
            now - timedelta(days=7)
        )

        thirty_days_start = (
            now - timedelta(days=30)
        )

        return {
            "total_solved": get_stat(
                statistics,
                "All",
            ),
            "easy": get_stat(
                statistics,
                "Easy",
            ),
            "medium": get_stat(
                statistics,
                "Medium",
            ),
            "hard": get_stat(
                statistics,
                "Hard",
            ),
            "submissions": get_stat(
                statistics,
                "All",
                "submissions",
            ),
            "solved_today":
                unique_solved_count_since(
                    recent_submissions,
                    today_start,
                ),
            "last_7_days":
                unique_solved_count_since(
                    recent_submissions,
                    seven_days_start,
                ),
            "last_30_days":
                unique_solved_count_since(
                    recent_submissions,
                    thirty_days_start,
                ),
            "last_problem": last_problem,
            "last_solved": last_solved,
            "status": "Success",
            "recent_submissions": recent_submissions,
        }

    except requests.Timeout:
        return empty_profile(
            "Request timeout"
        )

    except requests.ConnectionError as error:
        return empty_profile(
            f"Connection error: {error}"
        )

    except requests.RequestException as error:
        return empty_profile(
            f"Network error: {error}"
        )

    except Exception as error:
        return empty_profile(
            f"Unexpected error: {error}"
        )


# ============================================================
# SAFE CSV WRITE
# ============================================================

def atomic_csv_write(
    dataframe: pd.DataFrame,
    destination: Path,
) -> None:
    temporary_file = destination.with_suffix(
        ".temporary.csv"
    )

    dataframe.to_csv(
        temporary_file,
        index=False,
        encoding="utf-8-sig",
    )

    os.replace(
        temporary_file,
        destination,
    )


# ============================================================
# INPUT / SUPABASE
# ============================================================

def sync_students_from_supabase() -> None:
    """Download the authoritative student directory from Supabase."""

    if (
        not SUPABASE_URL
        or not SUPABASE_SERVICE_ROLE_KEY
    ):
        print(
            "Supabase secrets not set; "
            "using local students.csv"
        )
        return

    url = f"{SUPABASE_URL}/rest/v1/students"

    headers = {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization":
            f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Accept": "application/json",
    }

    params = {
        "select": (
            "register_number,"
            "student_name,"
            "leetcode_username,"
            "section,"
            "created_at"
        ),
        "order": "created_at.asc",
    }

    response = requests.get(
        url,
        headers=headers,
        params=params,
        timeout=30,
    )

    response.raise_for_status()

    rows = response.json()

    frame = pd.DataFrame(
        [
            {
                "Register Number":
                    clean(
                        row.get(
                            "register_number"
                        )
                    ),
                "Student Name":
                    clean(
                        row.get(
                            "student_name"
                        )
                    ),
                "LeetCode Username":
                    clean(
                        row.get(
                            "leetcode_username"
                        )
                    ),
                "Section":
                    clean(
                        row.get("section")
                    )
                    or "ECE E",
            }
            for row in rows
        ]
    )

    if frame.empty:
        frame = pd.DataFrame(
            columns=[
                "Register Number",
                "Student Name",
                "LeetCode Username",
                "Section",
            ]
        )

    atomic_csv_write(
        frame,
        STUDENTS_FILE,
    )

    print(
        f"Synced {len(frame)} student(s) "
        "from Supabase"
    )


def write_students_excel(
    students: pd.DataFrame,
) -> None:
    """Create an Excel copy automatically."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = [
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Section",
        "LeetCode Link",
    ]

    ws.append(headers)

    for _, student in students.iterrows():
        username = clean(
            student["LeetCode Username"]
        )

        ws.append(
            [
                clean(
                    student[
                        "Register Number"
                    ]
                ),
                clean(
                    student[
                        "Student Name"
                    ]
                ),
                username,
                clean(
                    student["Section"]
                ),
                (
                    "https://leetcode.com/u/"
                    f"{username}/"
                ),
            ]
        )

    header_fill = PatternFill(
        "solid",
        fgColor="2563EB",
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(
            color="FFFFFF",
            bold=True,
        )
        cell.alignment = Alignment(
            horizontal="center"
        )

    widths = {
        "A": 20,
        "B": 28,
        "C": 28,
        "D": 14,
        "E": 48,
    }

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(
        STUDENTS_XLSX
    )


def read_students() -> pd.DataFrame:
    sync_students_from_supabase()

    if not STUDENTS_FILE.exists():
        raise FileNotFoundError(
            f"students.csv not found: "
            f"{STUDENTS_FILE}"
        )

    students = pd.read_csv(
        STUDENTS_FILE,
        dtype=str,
        keep_default_na=False,
    )

    # Backward compatibility:
    # old local CSV files did not have Section.
    if "Section" not in students.columns:
        students["Section"] = "ECE E"

    required_columns = {
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Section",
    }

    missing = (
        required_columns
        .difference(students.columns)
    )

    if missing:
        raise ValueError(
            "Missing students.csv columns: "
            + ", ".join(
                sorted(missing)
            )
        )

    students = (
        students
        .dropna(how="all")
        .copy()
    )

    for column in required_columns:
        students[column] = (
            students[column]
            .apply(clean)
        )

    students.loc[
        students["Section"] == "",
        "Section",
    ] = "ECE E"

    invalid_sections = sorted(
        set(students["Section"])
        .difference(
            ALLOWED_SECTIONS
        )
    )

    if invalid_sections:
        raise ValueError(
            "Invalid section value(s): "
            + ", ".join(
                invalid_sections
            )
        )

    students = students[
        (
            students[
                "Student Name"
            ]
            != ""
        )
        & (
            students[
                "LeetCode Username"
            ]
            != ""
        )
    ].copy()

    write_students_excel(
        students
    )

    return students


# ============================================================
# COMPLETED-DAY ROLLING 7 / 30 DAY COUNTS
# ============================================================

def safe_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def load_daily_activity_file() -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
    ]

    if not DAILY_ACTIVITY_CSV.exists():
        return pd.DataFrame(columns=columns)

    try:
        frame = pd.read_csv(
            DAILY_ACTIVITY_CSV,
            dtype=str,
            keep_default_na=False,
        )
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=columns)

    for column in columns:
        if column not in frame.columns:
            frame[column] = ""

    return frame[columns].copy()


def latest_previous_snapshot(
    history: pd.DataFrame,
    register_number: str,
) -> dict[str, Any] | None:
    if history.empty:
        return None

    student_history = history[
        history["Register Number"].astype(str) == str(register_number)
    ].copy()

    if student_history.empty:
        return None

    student_history["_date"] = pd.to_datetime(
        student_history["Date"],
        errors="coerce",
    )

    today = pd.Timestamp(date.today())

    student_history = student_history[
        student_history["_date"] < today
    ].dropna(subset=["_date"])

    if student_history.empty:
        return None

    row = (
        student_history
        .sort_values("_date")
        .iloc[-1]
    )

    return row.to_dict()


def solved_on_date(
    activity: pd.DataFrame,
    register_number: str,
    target_date: date,
) -> int:
    if activity.empty:
        return 0

    matches = activity[
        (activity["Register Number"].astype(str) == str(register_number))
        & (activity["Date"].astype(str) == target_date.isoformat())
    ]

    if matches.empty:
        return 0

    return safe_int(matches.iloc[-1]["Solved That Day"])


def calculate_completed_day_counts(
    previous_history: pd.DataFrame,
    previous_activity: pd.DataFrame,
    register_number: str,
    current_total: int,
    solved_today: int,
    leetcode_7_days: int,
    leetcode_30_days: int,
) -> tuple[int, int, str, int]:
    """
    Keep the existing 7/30-day values, then update them once per new day.

    Today's solves are NOT added to 7/30 yet.
    On the next day they become the completed-day delta.

    Example:
      yesterday total = 100
      today total = 102
      solved_today = 0
      completed previous day = 2

    Those 2 are added to 7-day and 30-day counts today.
    """

    today = date.today()
    previous = latest_previous_snapshot(
        previous_history,
        register_number,
    )

    # First baseline: use LeetCode's current rolling values but exclude today.
    if previous is None:
        baseline_7 = max(0, safe_int(leetcode_7_days) - safe_int(solved_today))
        baseline_30 = max(0, safe_int(leetcode_30_days) - safe_int(solved_today))

        return (
            baseline_7,
            baseline_30,
            "",
            0,
        )

    previous_date_text = str(previous.get("Date", "")).strip()

    try:
        previous_date = date.fromisoformat(previous_date_text)
    except ValueError:
        previous_date = today - timedelta(days=1)

    previous_total = safe_int(
        previous.get("Problems Solved", 0)
    )

    previous_7 = safe_int(
        previous.get("Last 7 Days", 0)
    )

    previous_30 = safe_int(
        previous.get("Last 30 Days", 0)
    )

    # Total increase since the previous stored day, excluding today's solves.
    completed_delta = max(
        0,
        safe_int(current_total)
        - previous_total
        - safe_int(solved_today),
    )

    # We assign the completed increase to the day immediately before today.
    # With the normal daily/scheduled tracker this represents yesterday.
    completed_date = today - timedelta(days=1)

    expired_7 = solved_on_date(
        previous_activity,
        register_number,
        today - timedelta(days=7),
    )

    expired_30 = solved_on_date(
        previous_activity,
        register_number,
        today - timedelta(days=30),
    )

    last_7_days = max(
        0,
        previous_7 + completed_delta - expired_7,
    )

    last_30_days = max(
        0,
        previous_30 + completed_delta - expired_30,
    )

    return (
        last_7_days,
        last_30_days,
        completed_date.isoformat(),
        completed_delta,
    )


def update_completed_daily_activity(
    previous_activity: pd.DataFrame,
    current_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
    ]

    activity = previous_activity.copy()

    new_rows = []

    for row in current_rows:
        completed_date = row.get("_Completed Date", "")
        completed_solved = safe_int(
            row.get("_Completed Solved", 0)
        )

        if not completed_date:
            continue

        # Replace an existing record for this student/date.
        if not activity.empty:
            activity = activity[
                ~(
                    (activity["Date"].astype(str) == str(completed_date))
                    & (
                        activity["Register Number"].astype(str)
                        == str(row["Register Number"])
                    )
                )
            ].copy()

        new_rows.append({
            "Date": completed_date,
            "Section": row["Section"],
            "Register Number": row["Register Number"],
            "Student Name": row["Student Name"],
            "LeetCode Username": row["LeetCode Username"],
            "Problems Solved": row["Problems Solved"],
            "Solved That Day": completed_solved,
        })

    if new_rows:
        activity = pd.concat(
            [
                activity,
                pd.DataFrame(new_rows, columns=columns),
            ],
            ignore_index=True,
        )

    if not activity.empty:
        activity = activity.sort_values(
            by=["Date", "Section", "Register Number"],
            ascending=[True, True, True],
        ).reset_index(drop=True)

    return activity[columns]



# ============================================================
# DAILY CHALLENGE
# ============================================================

def supabase_headers() -> dict[str, str]:
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def load_recent_challenges(days: int = 35) -> list[dict[str, Any]]:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        return []

    start_date = (
        datetime.now(IST).date()
        - timedelta(days=days)
    ).isoformat()

    try:
        response = requests.get(
            f"{SUPABASE_URL}/rest/v1/daily_challenges",
            headers=supabase_headers(),
            params={
                "select": (
                    "id,challenge_date,problem_title,"
                    "problem_slug,problem_url,difficulty"
                ),
                "challenge_date": f"gte.{start_date}",
                "order": "challenge_date.asc",
            },
            timeout=30,
        )

        if not response.ok:
            print(
                "Daily challenge load warning: "
                f"{response.status_code} {response.text[:300]}"
            )
            return []

        return response.json()

    except requests.RequestException as error:
        print(
            "Daily challenge load network warning: "
            f"{error}"
        )
        return []


def challenge_completion(
    recent_submissions: list[dict[str, Any]],
    challenge: dict[str, Any],
) -> tuple[bool, str | None]:
    """
    Current rule:
    accepted submission must be on the challenge date in IST.
    """

    try:
        challenge_day = date.fromisoformat(
            str(challenge["challenge_date"])
        )
    except (KeyError, ValueError, TypeError):
        return False, None

    target_slug = clean(
        challenge.get("problem_slug")
    ).lower()

    if not target_slug:
        return False, None

    for submission in recent_submissions:
        if (
            clean(
                submission.get("titleSlug")
            ).lower()
            != target_slug
        ):
            continue

        timestamp = submission.get(
            "timestamp"
        )

        if not timestamp:
            continue

        try:
            solved_at = datetime.fromtimestamp(
                int(timestamp),
                tz=IST,
            )
        except (TypeError, ValueError, OSError):
            continue

        if solved_at.date() == challenge_day:
            return (
                True,
                solved_at.isoformat(),
            )

    return False, None


def save_challenge_result(
    challenge_id: int,
    register_number: str,
    completed: bool,
    completed_at: str | None,
) -> bool:
    """
    Robust challenge-result save:
    - GET then PATCH/POST (no on_conflict dependency)
    - Completed never becomes Pending
    - DB/network failures do not crash the LeetCode tracker
    """

    if (
        not SUPABASE_URL
        or not SUPABASE_SERVICE_ROLE_KEY
    ):
        return False

    register_number = clean(
        register_number
    )

    if not register_number:
        return False

    base_url = (
        f"{SUPABASE_URL}/rest/v1/"
        "daily_challenge_results"
    )

    headers = supabase_headers()

    try:
        lookup = requests.get(
            base_url,
            headers=headers,
            params={
                "select":
                    "id,completed,completed_at",
                "challenge_id":
                    f"eq.{challenge_id}",
                "register_number":
                    f"eq.{register_number}",
                "limit": "1",
            },
            timeout=30,
        )

        if not lookup.ok:
            print(
                "  Challenge lookup warning "
                f"[{register_number}]: "
                f"{lookup.status_code} "
                f"{lookup.text[:180]}"
            )
            return False

        existing_rows = lookup.json()

        existing_completed = False
        existing_completed_at = None

        if existing_rows:
            existing_completed = bool(
                existing_rows[0].get(
                    "completed"
                )
            )
            existing_completed_at = (
                existing_rows[0].get(
                    "completed_at"
                )
            )

        final_completed = (
            existing_completed
            or completed
        )

        final_completed_at = (
            existing_completed_at
            if existing_completed
            else completed_at
        )

        payload = {
            "completed":
                final_completed,
            "completed_at":
                final_completed_at,
            "checked_at":
                datetime.now(
                    IST
                ).isoformat(),
        }

        if existing_rows:
            response = requests.patch(
                base_url,
                headers={
                    **headers,
                    "Prefer":
                        "return=minimal",
                },
                params={
                    "challenge_id":
                        f"eq.{challenge_id}",
                    "register_number":
                        f"eq.{register_number}",
                },
                json=payload,
                timeout=30,
            )
        else:
            response = requests.post(
                base_url,
                headers={
                    **headers,
                    "Prefer":
                        "return=minimal",
                },
                json={
                    "challenge_id":
                        challenge_id,
                    "register_number":
                        register_number,
                    **payload,
                },
                timeout=30,
            )

        if not response.ok:
            print(
                "  Challenge save warning "
                f"[{register_number}]: "
                f"{response.status_code} "
                f"{response.text[:180]}"
            )
            return False

        return True

    except requests.RequestException as error:
        print(
            "  Challenge network warning "
            f"[{register_number}]: "
            f"{error}"
        )
        return False

    except Exception as error:
        print(
            "  Challenge unexpected warning "
            f"[{register_number}]: "
            f"{error}"
        )
        return False


# ============================================================
# PARALLEL STUDENT WORKER
# ============================================================

def process_student(
    position: int,
    total_students: int,
    student: pd.Series,
    updated_at: str,
    previous_history: pd.DataFrame,
    previous_activity: pd.DataFrame,
    recent_challenges: list[dict[str, Any]],
) -> dict[str, Any]:
    register_number = clean(
        student["Register Number"]
    )

    student_name = clean(
        student["Student Name"]
    )

    username = clean(
        student["LeetCode Username"]
    )

    section = clean(
        student["Section"]
    )

    print(
        f"[START {position}/{total_students}] "
        f"{section} | "
        f"{student_name} ({username})"
    )

    profile = fetch_leetcode(
        username
    )

    # Daily Challenge checks run inside this worker too.
    # Therefore LeetCode + challenge tracking both stay parallel.
    for challenge in recent_challenges:
        challenge_done, challenge_done_at = (
            challenge_completion(
                profile.get(
                    "recent_submissions",
                    [],
                ),
                challenge,
            )
        )

        challenge_saved = (
            save_challenge_result(
                int(
                    challenge["id"]
                ),
                register_number,
                challenge_done,
                challenge_done_at,
            )
        )

        if challenge_done:
            print(
                f"[CHALLENGE {position}/{total_students}] "
                f"{student_name} | "
                f"{challenge.get('problem_title', '')} | "
                f"{'saved ✅' if challenge_saved else 'save warning ⚠️'}"
            )

    (
        completed_7_days,
        completed_30_days,
        completed_date,
        completed_solved,
    ) = calculate_completed_day_counts(
        previous_history,
        previous_activity,
        register_number,
        profile["total_solved"],
        profile["solved_today"],
        profile["last_7_days"],
        profile["last_30_days"],
    )

    profile["last_7_days"] = completed_7_days
    profile["last_30_days"] = completed_30_days

    row = {
        "Section": section,
        "Register Number":
            register_number,
        "Student Name":
            student_name,
        "LeetCode Username":
            username,
        "LeetCode Link": (
            "https://leetcode.com/u/"
            f"{username}/"
        ),
        "Problems Solved":
            profile["total_solved"],
        "Solved Today":
            profile["solved_today"],
        "Last 7 Days":
            profile["last_7_days"],
        "Last 30 Days":
            profile["last_30_days"],
        "Total Submissions":
            profile["submissions"],
        "Easy":
            profile["easy"],
        "Medium":
            profile["medium"],
        "Hard":
            profile["hard"],
        "Last Problem":
            profile["last_problem"],
        "Last Solved":
            profile["last_solved"],
        "Status":
            profile["status"],
        "Updated At":
            updated_at,
        "_Completed Date":
            completed_date,
        "_Completed Solved":
            completed_solved,
    }

    print(
        f"[DONE  {position}/{total_students}] "
        f"{section} | "
        f"{student_name} | "
        f"30d={profile['last_30_days']} | "
        f"7d={profile['last_7_days']} | "
        f"today={profile['solved_today']} | "
        f"total={profile['total_solved']} | "
        f"{profile['status']}"
    )

    return row


# ============================================================
# RANKING
# ============================================================

def add_ranks(
    live_data: pd.DataFrame,
) -> pd.DataFrame:
    if live_data.empty:
        return live_data

    ranking_columns = [
        "Last 30 Days",
        "Last 7 Days",
        "Solved Today",
        "Problems Solved",
        "Student Name",
    ]

    ranking_ascending = [
        False,
        False,
        False,
        False,
        True,
    ]

    # --------------------------------------------------------
    # OVERALL RANK
    # --------------------------------------------------------

    overall_sorted = (
        live_data
        .sort_values(
            by=ranking_columns,
            ascending=ranking_ascending,
        )
        .reset_index(drop=True)
    )

    overall_sorted[
        "Overall Rank"
    ] = range(
        1,
        len(overall_sorted) + 1,
    )

    overall_rank_map = dict(
        zip(
            overall_sorted[
                "Register Number"
            ].astype(str),
            overall_sorted[
                "Overall Rank"
            ],
        )
    )

    live_data[
        "Overall Rank"
    ] = (
        live_data[
            "Register Number"
        ]
        .astype(str)
        .map(overall_rank_map)
    )

    # --------------------------------------------------------
    # SECTION RANK
    # --------------------------------------------------------

    section_rank_map: dict[
        tuple[str, str],
        int,
    ] = {}

    for (
        section,
        section_frame,
    ) in live_data.groupby(
        "Section",
        sort=False,
    ):
        section_sorted = (
            section_frame
            .sort_values(
                by=ranking_columns,
                ascending=ranking_ascending,
            )
            .reset_index(drop=True)
        )

        for index, row in (
            section_sorted
            .iterrows()
        ):
            key = (
                str(section),
                str(
                    row[
                        "Register Number"
                    ]
                ),
            )

            section_rank_map[
                key
            ] = index + 1

    live_data[
        "Section Rank"
    ] = live_data.apply(
        lambda row:
            section_rank_map.get(
                (
                    str(
                        row["Section"]
                    ),
                    str(
                        row[
                            "Register Number"
                        ]
                    ),
                ),
                "",
            ),
        axis=1,
    )

    # Store CSV in overall-rank order.
    live_data = (
        live_data
        .sort_values(
            by=[
                "Overall Rank"
            ],
            ascending=[True],
        )
        .reset_index(drop=True)
    )

    columns = [
        "Overall Rank",
        "Section Rank",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "LeetCode Link",
        "Problems Solved",
        "Solved Today",
        "Last 7 Days",
        "Last 30 Days",
        "Total Submissions",
        "Easy",
        "Medium",
        "Hard",
        "Last Problem",
        "Last Solved",
        "Status",
        "Updated At",
    ]

    return live_data[columns]


# ============================================================
# HISTORY
# ============================================================

HISTORY_COLUMNS = [
    "Date",
    "Overall Rank",
    "Section Rank",
    "Section",
    "Register Number",
    "Student Name",
    "LeetCode Username",
    "Problems Solved",
    "Solved Today",
    "Last 7 Days",
    "Last 30 Days",
    "Total Submissions",
    "Easy",
    "Medium",
    "Hard",
    "Last Problem",
    "Last Solved",
    "Status",
    "Updated At",
]


def load_history() -> pd.DataFrame:
    if not HISTORY_CSV.exists():
        return pd.DataFrame(
            columns=HISTORY_COLUMNS
        )

    try:
        history = pd.read_csv(
            HISTORY_CSV,
            dtype=str,
            keep_default_na=False,
        )
    except pd.errors.EmptyDataError:
        return pd.DataFrame(
            columns=HISTORY_COLUMNS
        )

    for column in HISTORY_COLUMNS:
        if column not in history.columns:
            history[column] = ""

    return history[
        HISTORY_COLUMNS
    ].copy()


def update_history(
    previous_history: pd.DataFrame,
    current_rows: list[
        dict[str, Any]
    ],
) -> pd.DataFrame:
    today_text = (
        date.today().isoformat()
    )

    history = (
        previous_history.copy()
    )

    if not history.empty:
        current_registers = {
            str(
                row[
                    "Register Number"
                ]
            )
            for row in current_rows
        }

        history = history[
            ~(
                (
                    history["Date"]
                    == today_text
                )
                & (
                    history[
                        "Register Number"
                    ]
                    .astype(str)
                    .isin(
                        current_registers
                    )
                )
            )
        ].copy()

    new_history_rows = []

    for row in current_rows:
        new_history_rows.append(
            {
                "Date":
                    today_text,
                "Overall Rank":
                    row.get(
                        "Overall Rank",
                        "",
                    ),
                "Section Rank":
                    row.get(
                        "Section Rank",
                        "",
                    ),
                "Section":
                    row.get(
                        "Section",
                        "",
                    ),
                "Register Number":
                    row[
                        "Register Number"
                    ],
                "Student Name":
                    row[
                        "Student Name"
                    ],
                "LeetCode Username":
                    row[
                        "LeetCode Username"
                    ],
                "Problems Solved":
                    row[
                        "Problems Solved"
                    ],
                "Solved Today":
                    row[
                        "Solved Today"
                    ],
                "Last 7 Days":
                    row[
                        "Last 7 Days"
                    ],
                "Last 30 Days":
                    row[
                        "Last 30 Days"
                    ],
                "Total Submissions":
                    row[
                        "Total Submissions"
                    ],
                "Easy":
                    row["Easy"],
                "Medium":
                    row["Medium"],
                "Hard":
                    row["Hard"],
                "Last Problem":
                    row[
                        "Last Problem"
                    ],
                "Last Solved":
                    row[
                        "Last Solved"
                    ],
                "Status":
                    row["Status"],
                "Updated At":
                    row[
                        "Updated At"
                    ],
            }
        )

    combined = pd.concat(
        [
            history,
            pd.DataFrame(
                new_history_rows,
                columns=HISTORY_COLUMNS,
            ),
        ],
        ignore_index=True,
    )

    if not combined.empty:
        combined = (
            combined
            .sort_values(
                by=[
                    "Date",
                    "Section",
                    "Student Name",
                ],
                ascending=[
                    True,
                    True,
                    True,
                ],
            )
            .reset_index(
                drop=True
            )
        )

    return combined


def build_daily_activity(
    history: pd.DataFrame,
) -> pd.DataFrame:
    columns = [
        "Date",
        "Section",
        "Register Number",
        "Student Name",
        "LeetCode Username",
        "Problems Solved",
        "Solved That Day",
    ]

    if history.empty:
        return pd.DataFrame(
            columns=columns
        )

    activity = history[
        [
            "Date",
            "Section",
            "Register Number",
            "Student Name",
            "LeetCode Username",
            "Problems Solved",
            "Solved Today",
        ]
    ].copy()

    activity = activity.rename(
        columns={
            "Solved Today":
                "Solved That Day"
        }
    )

    return activity[
        columns
    ]


# ============================================================
# MAIN UPDATE
# ============================================================

def run_one_update() -> None:
    students = read_students()

    updated_at = datetime.now(IST).strftime(
        "%Y-%m-%d %H:%M:%S IST"
    )

    live_rows: list[
        dict[str, Any]
    ] = []

    total_students = len(
        students
    )

    worker_count = (
        min(
            MAX_WORKERS,
            total_students,
        )
        if total_students
        else 1
    )

    print("=" * 64)
    print(
        f"LeetCode cloud update: "
        f"{updated_at}"
    )
    print(
        f"Students: "
        f"{total_students}"
    )
    # Load shared data BEFORE printing/using it.
    previous_history = load_history()
    previous_activity = load_daily_activity_file()

    # Load challenge definitions once.
    # The same read-only list is shared safely across all workers.
    recent_challenges = load_recent_challenges()

    print(
        f"Parallel workers: "
        f"{worker_count}"
    )
    print(
        f"Daily challenges loaded: "
        f"{len(recent_challenges)}"
    )
    print("=" * 64)

    futures = {}

    with ThreadPoolExecutor(
        max_workers=worker_count
    ) as executor:
        for (
            position,
            (_, student),
        ) in enumerate(
            students.iterrows(),
            start=1,
        ):
            future = executor.submit(
                process_student,
                position,
                total_students,
                student,
                updated_at,
                previous_history,
                previous_activity,
                recent_challenges,
            )

            futures[
                future
            ] = {
                "position":
                    position,
                "section":
                    clean(
                        student[
                            "Section"
                        ]
                    ),
                "student_name":
                    clean(
                        student[
                            "Student Name"
                        ]
                    ),
                "register_number":
                    clean(
                        student[
                            "Register Number"
                        ]
                    ),
                "username":
                    clean(
                        student[
                            "LeetCode Username"
                        ]
                    ),
            }

        for future in as_completed(
            futures
        ):
            info = futures[
                future
            ]

            try:
                live_rows.append(
                    future.result()
                )

            except Exception as error:
                print(
                    f"[FAILED "
                    f"{info['position']}/"
                    f"{total_students}] "
                    f"{info['section']} | "
                    f"{info['student_name']} "
                    f"({info['username']}): "
                    f"{error}"
                )

                live_rows.append(
                    {
                        "Section":
                            info[
                                "section"
                            ],
                        "Register Number":
                            info[
                                "register_number"
                            ],
                        "Student Name":
                            info[
                                "student_name"
                            ],
                        "LeetCode Username":
                            info[
                                "username"
                            ],
                        "LeetCode Link":
                            (
                                "https://leetcode.com/u/"
                                f"{info['username']}/"
                            ),
                        "Problems Solved":
                            0,
                        "Solved Today":
                            0,
                        "Last 7 Days":
                            0,
                        "Last 30 Days":
                            0,
                        "Total Submissions":
                            0,
                        "Easy":
                            0,
                        "Medium":
                            0,
                        "Hard":
                            0,
                        "Last Problem":
                            "",
                        "Last Solved":
                            "",
                        "Status":
                            (
                                "Worker error: "
                                f"{error}"
                            ),
                        "Updated At":
                            updated_at,
                        "_Completed Date":
                            "",
                        "_Completed Solved":
                            0,
                    }
                )

    # Keep the internal completed-day data for DailyActivity.csv.
    completed_activity_rows = list(live_rows)

    live_data = pd.DataFrame(
        live_rows
    )

    # Internal helper columns must not appear in LiveData.csv.
    live_data = live_data.drop(
        columns=[
            "_Completed Date",
            "_Completed Solved",
        ],
        errors="ignore",
    )

    live_data = add_ranks(
        live_data
    )

    history_rows = (
        live_data.to_dict(
            "records"
        )
        if not live_data.empty
        else []
    )

    history = update_history(
        previous_history,
        history_rows,
    )

    daily_activity = update_completed_daily_activity(
        previous_activity,
        completed_activity_rows,
    )

    atomic_csv_write(
        live_data,
        LIVE_CSV,
    )

    atomic_csv_write(
        history,
        HISTORY_CSV,
    )

    atomic_csv_write(
        daily_activity,
        DAILY_ACTIVITY_CSV,
    )

    print("=" * 64)
    print(
        "CSV files updated "
        "successfully."
    )
    print(
        f"LiveData.csv: "
        f"{LIVE_CSV}"
    )
    print(
        f"History.csv: "
        f"{HISTORY_CSV}"
    )
    print(
        f"DailyActivity.csv: "
        f"{DAILY_ACTIVITY_CSV}"
    )
    print("=" * 64)


if __name__ == "__main__":
    run_one_update()
